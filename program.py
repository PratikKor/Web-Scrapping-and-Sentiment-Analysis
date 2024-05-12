import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from nltk.tokenize import word_tokenize, sent_tokenize    
import re
import os
    
#primary statement for sucesfull launch of program
print("Fetching content from webpages...")

#scrapping data from webpages
def fetch_content(url, URL_ID):
    page = requests.get(url)
    
    if page.status_code == 200:
        soup = BeautifulSoup(page.content, 'html.parser')
        
        #getting the respective Title and Article text from webpages
        title = soup.find("h1", {'class': 'entry-title'}) or soup.find("h1", {'class': 'tdb-title-text'})
        article = soup.find(attrs={'class': 'td-post-content tagdiv-type'}) or soup.find(attrs={'class': 'tdb-block-inner td-fix-index'})
        
        #replacing the new lines with whitespace and creating .txt files with respective names form URL_ID
        if title and article:
            title_text = title.text.replace('\n', ' ')
            article_text = article.text.replace('\n', ' ')
            
            file_name = f"txt_files/{URL_ID}.txt"
            with open(file_name, "w") as file:
                file.write(title_text)
                file.write(article_text)
            print(f"File created for {URL_ID}")
            
        #case handling for empty webpage / data from webpage    
        else:
            print(f"Unable to extract title or article from {URL_ID}")
            create_empty_file(URL_ID, url)
    else:
        print(f"Failed to fetch content from {URL_ID}")
        create_empty_file(URL_ID, url)

def create_empty_file(URL_ID, url):
    file_name = f"txt_files/{URL_ID}.txt"
    with open(file_name, "w") as file:
        file.write(" ")
    print(f"Empty file created for {URL_ID}")

df = pd.read_excel('Input.xlsx')

#making txt_files folder to store all the .txt files generated
os.makedirs('txt_files')

for index, row in df.iterrows():
    url = row['URL']
    URL_ID = row['URL_ID']
    fetch_content(url, URL_ID)

print("Data Fetched from Webpages")

print("Processing files...")

#Analysis and adding data to excel
def load_words_from_file(file_path):
    with open(file_path, 'r', encoding='latin-1') as file:
        return set(file.read().splitlines())

def clean_text(text, stop_words):
    return [word.lower() for word in word_tokenize(text) if word.isalpha() and word.lower() not in stop_words]

def calculate_sentiment_scores(cleaned_words, positive_words, negative_words):
    positive_score = sum(1 for word in cleaned_words if word in positive_words)
    negative_score = sum(1 for word in cleaned_words if word in negative_words)
    total_score = positive_score + negative_score
    
    subjectivity_score = total_score / len(cleaned_words) if cleaned_words else 0.0
    polarity_score = (positive_score - negative_score) / total_score if total_score else 0.0
    
    return positive_score, negative_score, round(polarity_score, 3), round(subjectivity_score, 3)

def calculate_readability_metrics(text, stop_words):
    words = clean_text(text, stop_words)
    sentences = sent_tokenize(text)
    if not sentences:
        return [0] * 9 
    complex_words = [word for word in words if len(word) > 2]
    avg_sentence_length = len(words) / len(sentences)
    percentage_complex_words = len(complex_words) / len(words)
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)
    avg_words_per_sentence = len(words) / len(sentences)
    return [
        round(avg_sentence_length, 3), round(percentage_complex_words, 3), round(fog_index, 3),
        round(avg_words_per_sentence, 3), len(complex_words), len(words),
        sum(len(re.findall('[aeiou]+', word)) for word in words),
        sum(1 for word in words if re.match(r'\b(?:I|we|my|ours|us)\b', word, flags=re.IGNORECASE)),
        round(sum(len(word) for word in words) / len(words), 3)
    ]

#stop words
stop_words = set()
stop_words_path = 'StopWords'
for filename in os.listdir(stop_words_path):
    if filename.endswith('.txt'):
        stop_words.update(load_words_from_file(os.path.join(stop_words_path, filename)))
print("Stop words loaded.")

#positive & negative words
positive_words = load_words_from_file('MasterDictionary/positive-words.txt')
negative_words = load_words_from_file('MasterDictionary/negative-words.txt')
print("Positive and negative dictionaries loaded.")

#To Load Excel File
print("Loading Excel file...")
file_path = "Output Data Structure.xlsx"
wb = load_workbook(filename=file_path)
sheet = wb.active

last_row_index = sheet.max_row


#Looping through each file in txt_file Folder
folder_path = 'txt_files'
for filename in sorted(os.listdir(folder_path)):
    if filename.endswith('.txt'):
        #removing the .txt extension so :-4
        url_id = filename[:-4]
        with open(os.path.join(folder_path, filename), 'r', encoding='latin-1') as file:
            text_to_analyze = file.read()
        
        # Perform sentiment analysis and readability calculations
        cleaned_words = clean_text(text_to_analyze, stop_words)
        sentiment_scores = calculate_sentiment_scores(cleaned_words, positive_words, negative_words)
        readability_metrics = calculate_readability_metrics(text_to_analyze, stop_words)
        
        #finding the respective row with URL_ID
        found_row_index = next((row_index for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=last_row_index, min_col=1, max_col=1, values_only=True), start=2) if row[0] == url_id), None)
        
        if found_row_index is not None:
            #Updating the columns
            for i, value in enumerate(sentiment_scores + tuple(readability_metrics), start=3):
                sheet.cell(row=found_row_index, column=i, value=round(value, 3))

print("Saving Excel file...")
wb.save(file_path)
print("Excel file saved.")

